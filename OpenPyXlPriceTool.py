
#program automates process of creating spread sheet and instituting price changes
#uses excel conditional formating to make easy identification or price changes for the user
#Version1  010419


import os
import time
from openpyxl import Workbook,load_workbook




def SheetBuilder(bookName):
    wb = Workbook()
    ws = wb.active

    #ToDo: rewrite to allow user to define sheet headers and pricing in gui
    
    data = [
        ["Crown Royal Flavor", "Season","MSRP","Sale Price","PriceChange T,F"],
        ["Peach", "Summer",21.99],
        ["Vanilla", "Winter",21.99],
        ["Cotton Candy", "Summer",25.99],
        ["Rose Petals", "Spring",150.00],
        ]

    for r in data:
        ws.append(r)
    wb.save(bookName)

    
    

def PriceChanger(bookName):
    wb = load_workbook(bookName)
    ws = wb.active

    targetPrice = 25

    ws["D2"]="=IF(C2>25,C2-(C2*.05),C2)"
    ws["D3"]="=IF(C3>25,C3-(C3*.05),C3)"
    ws["D4"]="=IF(C4>25,C4-(C4*.05),C4)"
    ws["D5"]="=IF(C5>25,C5-(C5*.05),C5)"
    #ws["D6"]="=IF(C6>25,C6-(C6*.05),C6)"


    #ToDo: create function that allows user to determine range of cells needed
    #ToDo: function to include discount percentage
    wb.save(bookName)

    os.startfile(bookName)#opens excel book for user.
  
bookName = "openpyXLbackend.xlsx" #excel file name
        
SheetBuilder(bookName)
PriceChanger(bookName)

